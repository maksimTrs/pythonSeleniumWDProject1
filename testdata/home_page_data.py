import openpyxl


class HomePageData:
    test_HomePage_data = [{"firstname": "Rahul", "lastname": "shetty", "gender": "Male", "birthday": "1995-05-25"},
                          {"firstname": "Anshika", "lastname": "shetty", "gender": "Female", "birthday": "2020-02-29"}]

    @staticmethod
    def getTestData(test_case_name):
        Dict = {}
        book = openpyxl.load_workbook("E:/MAX/IT/UDEMY/Selenium_Webdriver_with_PYTHON_from "
                                      "Scratch_Frameworks/PythonDemo.xlsx")
        sheet = book.active
        for i in range(1, sheet.max_row + 1):  # to get rows
            if sheet.cell(row=i, column=1).value == test_case_name:

                for j in range(2, sheet.max_column + 1):  # to get columns
                    # Dict["lastname"]="shetty
                    Dict[sheet.cell(row=1, column=j).value] = sheet.cell(row=i, column=j).value
        return [Dict]
